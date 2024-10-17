import os
import random
import pprint
import openpyxl
import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime, timedelta
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from data import courses, short_teachers, shortsub,CLASSES,n_m,m,ele,ele_sub,list1,list2,list3,CLASSES_t,cc



class TimetableGenerator:
    def __init__(self, courses, num_days=6, num_hours=9):
        self.courses = courses
        self.classes = list(CLASSES.keys())
        self.ct=list(CLASSES_t.keys())
        self.num_days = num_days
        self.num_hours = num_hours
        self.notmorningteachers=[]
        self.lab_timetables = {
            "BCA Lab": {day: {hour: [] for hour in range(num_hours)} for day in range(num_days)},
            "MCA Lab": {day: {hour: [] for hour in range(num_hours)} for day in range(num_days)},
            "BSc Lab": {day: {hour: [] for hour in range(num_hours)} for day in range(num_days)}
        }
        for h in range(self.num_hours):
            self.lab_timetables["MCA Lab"][5][h].append(("NULL", "NULL"))
        self.teacher_schedule = {day: {hour: set() for hour in range(num_hours)} for day in range(num_days)}  # Initialize teacher_schedule
        self.labs_assigned_per_hour = {day: {hour: 0 for hour in range(self.num_hours)} for day in range(self.num_days)}
        self.lab_teachers = {class_name: {subject: [] for category in categories.values() for subject in category.keys()}
                for class_name, categories in courses.items()}        
        self.timetable = None  # Initialize timetable as None
    def generate_timetable(self):
        
        
        # Initialize timetable structure
        timetable = self.initialize_timetable()
        # pprint.pprint(timetable)
        allocated_subjects, theory_hours_assigned, lab_assigned_per_day, teacher_schedule = self.initialize_tracking_structures()
        self.blockteachers(self.notmorningteachers)
        # Pre-assign subjects
        self.pre_assign_subjects(timetable)
        blocklib=None
        temp=self.find_free_slots_closest_to_six(timetable)
        for i,j in temp.items():
            if i==5:
                continue
            if 5 in j:
                blocklib=i
        self.pre_assign_extra_subjects(timetable,blocklib)
        
       
      

        
         # Assign LCA subjects to the first two hours of the day
        self.assign_lca_and_lab_hours(timetable, allocated_subjects, lab_assigned_per_day)
        
  
           #  # Assign all elective subjects at the same time for both sections A and B
        
        
        self.assign_elective_lab(timetable, allocated_subjects, lab_assigned_per_day, teacher_schedule)       
        self.assign_elective_hours(timetable, allocated_subjects, teacher_schedule)
        
        self.assign_lab_hours_multiple_electives(timetable, allocated_subjects, lab_assigned_per_day, teacher_schedule, "5CME")
        self.assign_elective_hours_single_section(timetable, allocated_subjects, teacher_schedule, "5CME")

        # Assign lab hours and theory hours
  
        self.assign_lab_hours(timetable, allocated_subjects, lab_assigned_per_day, teacher_schedule)
        self.assign_theory_hours(timetable, allocated_subjects, theory_hours_assigned, teacher_schedule)
        
        self.assign_elective_teacher(timetable)
        self.assign_teachers_to_all_labs(timetable)  
        
           
        
        self.print_teacher_hr(self.calculate_teacher_hours())
      
        
        
        self.cross_verify_timetable_with_courses(timetable)
        
        # pprint.pprint(self.lab_teachers)
        
       
        
        
        self.timetable = timetable
        
        # pprint.pprint(timetable)

        return timetable
    
    def blockteachers(self,notmorningteachers):
        for i in notmorningteachers:
            for d in range(self.num_days):
                for hr in range(0,2):
                    self.teacher_schedule[d][hr].add(i)  # Block the teacher's schedule

       
                
    
    def find_free_slots_closest_to_six(self, timetable):
        """
        Traverse the timetable and find all days and hours where all classes are free,
        then return the hours closest to hour index 6.

        :param timetable: A dictionary representing the timetable where 
                        timetable[day][class_name][hour] gives the schedule.
        :return: A dictionary where keys are days and values are sets of hours closest to 6 that are free.
        """
        free_slots = {day: [] for day in range(self.num_days)}  # Initialize free slots for each day

        # Traverse each day in the timetable
        for day in range(self.num_days):
            # Check each hour for all classes
            for hour in range(self.num_hours):
                hour_free = True  # Assume the hour is free initially

                # Check all classes for the current hour
                for class_name in self.classes:  # Assuming `self.classes` contains all class names
                    if timetable[day][class_name][hour]:  # If there is any booking in this hour
                        hour_free = False  # This hour is occupied
                        break  # No need to check further for this hour

                # If the hour is free for all classes, record it
                if hour_free:
                    free_slots[day].append(hour)

        # Now find the hours closest to 6 for each day
        # closest_hours = {}
        # for day, hours in free_slots.items():
        #     # Calculate distances to hour 6 and find the closest hour(s)
        #     closest_to_six = {hour for hour in hours if abs(hour - 6) == min(abs(hour - 6) for hour in hours)}
        #     closest_hours[day] = closest_to_six

        # Return the dictionary of closest hours for each day
        return free_slots

    

    
   

    
    def assign_lca_and_lab_hours(self, timetable, allocated_subjects, lab_assigned_per_day):
        for class_name in self.classes:
            if 'LCA' in self.courses[class_name]:
                lca_subjects = self.courses[class_name]['LCA']
                for subject_name, subject_info in lca_subjects.items():
                    lab_hours = subject_info["lab_hours"]
                    teachers_for_subject = subject_info["teacher_incharge"]
                    assigned_lab_hours = 0

                    for day in range(self.num_days):
                        if assigned_lab_hours >= lab_hours:
                            break

                        if not lab_assigned_per_day[day][class_name]:
                            if (len(timetable[day][class_name][0]) == 0 and len(timetable[day][class_name][1]) == 0 and
                                    self.labs_assigned_per_hour[day][0] ==0):  # Check if less than 2 labs assigned

                                p = self.is_teacher_allocated_in_first_two_hours(teachers_for_subject, day, 0)
                                if all(teacher not in self.teacher_schedule[day][0] for teacher in teachers_for_subject) and p:
                                    lab_venue = self.get_available_lab(day, 0)
                                    if lab_venue and self.is_lab_available(lab_venue, day, 0):
                                        self.assign_lab(timetable, day, class_name, 0, subject_name, teachers_for_subject, lab_venue, allocated_subjects, self.teacher_schedule)
                                        
                                        lab_assigned_per_day[day][class_name] = True
                                        self.labs_assigned_per_hour[day][0] += 1
                                        self.labs_assigned_per_hour[day][1] += 1
                                                                                    # Increment labs count for that hour
                                        assigned_lab_hours += 1
                                        break

                    # Now, handle the normal LCA hours
                    required_hours = subject_info["normal_hours"]  # Required LCA hours
                    assigned_lca_hours = 0

                    # Assign LCA hours incrementally
                    for day in range(self.num_days):
                        if assigned_lca_hours >= required_hours:
                            break  # Stop if all required LCA hours have been assigned

                        # Check if we can assign two consecutive hours in the first two hours of the day
                        for hour in range(2):  # Only check hours 0 and 1
                            teachers_for_lca = subject_info["teacher_incharge"]
                            # Ensure both hours are free and teachers are available
                            can_assign = (len(timetable[day][class_name][hour]) == 0 and 
                                        len(timetable[day][class_name][hour + 1]) == 0 and
                                        all(teacher not in self.teacher_schedule[day][hour] and teacher not in self.teacher_schedule[day][hour + 1]
                                            for teacher in teachers_for_lca))
                            
                            p=self.is_teacher_allocated_in_first_two_hours(teachers_for_lca,day,hour)  and self.is_teacher_allocated_in_first_two_hours(teachers_for_lca,day,hour)

                            # Try to assign hours one by one to meet the required total
                            if can_assign:
                                if assigned_lca_hours + 1 <= required_hours:
                                    # Assign the first hour
                                    timetable[day][class_name][hour].append((subject_name, teachers_for_lca))
                                    for teacher in teachers_for_lca:
                                        self.teacher_schedule[day][hour].add(teacher)  # Block the teacher's schedule
                                    assigned_lca_hours += 1  # Increment the assigned LCA hours

                                if assigned_lca_hours < required_hours and assigned_lca_hours + 1 <= required_hours:
                                    # Assign the second hour
                                    timetable[day][class_name][hour + 1].append((subject_name, teachers_for_lca))
                                    for teacher in teachers_for_lca:
                                        self.teacher_schedule[day][hour + 1].add(teacher)  # Block the teacher's schedule
                                    assigned_lca_hours += 1  # Increment again

                                break  # Move to the next day after assigning hours

                        # If the LCA subject has been assigned for all required hours, stop assigning
                        if assigned_lca_hours >= required_hours:
                            break


    
        
       
    
    def assign_elective_teacher(self, timetable):
        """
        Assigns a total of 5 teachers to all lab sessions, updates the timetable,
        and blocks the teacher's schedule for all relevant days.

        Args:
        timetable (list): The timetable structure.
        """
        # Get all lab hours available in the timetable
        lab_hours = self.get_lab_hours(timetable)
        
        # Teacher's maximum lab hours constraints
        max_teaching_hours = 18
        sagaya_hours_limit = 12
        manjunatha_hours_limit = 12
        max_teachers = 5  

        # Iterate over each lab session
        for (class_name, subject_name), lab_time_slots in lab_hours.items():
            # Skip if the subject matches the elective condition
            if (class_name == ele[1] and subject_name in ele_sub[1]) or\
                (class_name == ele[0] and subject_name in ele_sub[0]) or \
                (class_name=="5CME" and subject_name in ["WEB TECHNOLOGY","GRAPHICS AND ANIMATION"]):
                # Prepare the lab time slots based on the elective class
                if class_name == ele[1]:
                    lab_time_slots_set = set(lab_time_slots[0:2])  # Assume elective 1 has slots 2 to 4
                    lab_time_slots = lab_time_slots[0:2]
                elif class_name == ele[0]:
                    lab_time_slots_set = set(lab_time_slots[2:5])  # Assume elective 0 has first 2 slots
                    lab_time_slots = lab_time_slots[2:5]
                elif class_name=="5CME":
                    lab_time_slots_set = set(lab_time_slots[0:2])  # Assume elective 1 has slots 2 to 4
                    lab_time_slots = lab_time_slots[0:2]
                    
                

                # List to keep track of allocated teachers for this lab
                allocated_teachers = []

                # Check existing teachers assigned to the lab session
                for (day, hour) in lab_time_slots:
                    if len(timetable[day][class_name][hour]) > 0:
                        if class_name == ele[1]:
                            allocated_teachers.extend(timetable[day][class_name][hour][0][1])  # Teachers at index 1 for elective 1
                        elif class_name == ele[0]:
                            allocated_teachers.extend(timetable[day][class_name][hour][1][1])  # Teachers at index 1 for elective 0
                        elif class_name=="5CME":
                            if subject_name=="WEB TECHNOLOGY":
                                allocated_teachers.extend(timetable[day][class_name][hour][0][1])  # Teachers at index 1 for elective 0
                            elif subject_name=="GRAPHICS AND ANIMATION":
                                allocated_teachers.extend(timetable[day][class_name][hour][1][1])
                            
                        # Stop if the number of allocated teachers exceeds the limit
                        if len(allocated_teachers) >= max_teachers:
                            break

                # Ensure unique teachers
                allocated_teachers = list(set(allocated_teachers))
                teacher_free_hours = self.get_teacher_free_hours(self.num_days, self.num_hours)
                teacher_hours = self.calculate_teacher_hours()

                # Check each teacher's availability
                for teacher, free_slots in teacher_free_hours.items():
                    free_slots_set = set(free_slots)

                    # Check if specific teachers have exceeded their hourly limits
                    if (teacher.lower() == "dr sagaya aurelia p".lower() and teacher_hours.get(teacher, 0) >= sagaya_hours_limit) or \
                    (teacher.lower() == "dr manjunatha hiremath".lower() and teacher_hours.get(teacher, 0) >= manjunatha_hours_limit):
                        continue

                    # Check if the lab hours are a subset of the teacher's free hours
                    if lab_time_slots_set.issubset(free_slots_set):
                        current_hours = teacher_hours.get(teacher, 0)
                        first_two_hours_allocated = any(teacher in self.teacher_schedule[day][h] for h in range(0,2))
                        last_two_hours_allocated = any(teacher in self.teacher_schedule[day][h] for h in range(7, 9))
                        lab_last2=any(h in range(7, 9) for _, h in lab_time_slots)
                        lab_first2=any(h in range(0,2) for _, h in lab_time_slots)

                        # If the teacher is assigned to the first two hours, do not assign to the last two hours and vice versa
                        if first_two_hours_allocated and lab_last2:
                            
                            continue  # Skip assigning teacher if there's a conflict with last two hours
                        if last_two_hours_allocated and lab_first2:
                            continue  # Skip assigning teacher if there's a conflict with first two hours


                        # Check if adding the current lab session will exceed max teaching hours
                        if current_hours + 2 <= max_teaching_hours:
                            # Allocate the teacher
                            allocated_teachers.append(teacher)
                            teacher_hours[teacher] = current_hours + 2

                            # Stop allocating if we have reached the required number of teachers
                            if len(allocated_teachers) >= max_teachers:
                                break

                # Ensure we only keep unique teachers again
                allocated_teachers = list(set(allocated_teachers))
                if class_name in ele:
                    self.lab_teachers[ele[0]][subject_name].append(allocated_teachers)
                    self.lab_teachers[ele[1]][subject_name].append(allocated_teachers)
                else:
                    self.lab_teachers[class_name][subject_name].append(allocated_teachers)
                    
                    
                # Block the teacher's schedule for the lab time slots
                for (day, hour) in lab_time_slots:
                    if hour <= 8:  # Ensure valid hours
                        if allocated_teachers:  # Only update if we have allocated teachers
                            # Update the timetable with the allocated teachers
                            if class_name in ele:
                                # print(class_name)
                                timetable[day][ele[0]][hour].append(("", allocated_teachers))  # Add to the timetable
                                timetable[day][ele[1]][hour].append(("", allocated_teachers))
                               
                            else:
                                timetable[day][class_name][hour].append(("", allocated_teachers))

                            # Block the teacher's schedule for all relevant days and hours
                            for teacher in allocated_teachers:
                                self.teacher_schedule[day][hour].add(teacher)  # Ensure only teacher names are added

                # print(f"{class_name}: Allocated Teachers -> {allocated_teachers}")
                
                # If fewer than 6 teachers are allocated, notify
                if len(allocated_teachers) < max_teachers:
                    print(f"Only {len(allocated_teachers)} teachers allocated for {class_name} - {subject_name} lab. Need more teachers.")

    def assign_teachers_to_all_labs(self, timetable):
        """
        Assigns a total of 5 teachers to all lab sessions, updates the timetable,
        and blocks the teacher's schedule for all relevant days.

        Args:
        timetable (list): The timetable structure.
        """
        # Get all lab hours available in the timetable
        lab_hours = self.get_lab_hours(timetable)

        # Iterate over each lab session
        for (class_name, subject_name), lab_time_slots in lab_hours.items():
            # Skip elective subjects
            if subject_name not in ele_sub and not (class_name=="5CME" and subject_name in ["WEB TECHNOLOGY","GRAPHICS AND ANIMATION"]):
                # Convert lab_time_slots to a set for easier subset checking
                if len(set(lab_time_slots))==2:
                    lab_time_slots_set = set(lab_time_slots[0:2])
                    lab_time_slots=lab_time_slots[0:2]
                else:
                    lab_time_slots_set = set(lab_time_slots[0:2])
                    lab_time_slots=lab_time_slots[0:2]

                # List to track allocated teachers for this lab session
                allocated_teachers = []

                # Check existing teachers already assigned to the lab session
                for (day, hour) in lab_time_slots:
                    if len(timetable[day][class_name][hour]) > 0:
                        # Get pre-assigned teachers (assumes teachers are at index 1)
                        allocated_teachers.extend(timetable[day][class_name][hour][0][1])

                # Ensure we have unique teachers initially
                allocated_teachers = list(set(allocated_teachers))

                # Get teacher availability and calculate teaching hours
                teacher_free_hours = self.get_teacher_free_hours(self.num_days, self.num_hours)
                teacher_hours = self.calculate_teacher_hours()

                # Iterate through available teachers to fill the remaining slots
                for teacher, free_slots in teacher_free_hours.items():
                    free_slots_set = set(free_slots)

                    # Check teacher-specific hour limits (e.g., max 12 hours for specific teachers)
                    if ((teacher.lower() == "dr sagaya aurelia p".lower() and teacher_hours.get(teacher, 0) >= 12) or 
                        (teacher.lower() == "dr manjunatha hiremath".lower() and teacher_hours.get(teacher, 0) >= 12)):
                        continue  # Skip this teacher if they've reached their limit

                    # Ensure the teacher has enough free slots for the lab
                    if lab_time_slots_set.issubset(free_slots_set):
                        current_hours = teacher_hours.get(teacher, 0)

                        # Ensure adding this lab won't exceed 18 hours for the teacher
                        if current_hours + 2 <= 18:
                            # Add checks for the first 2 and last 2 hours condition
                            first_two_hours_allocated = any(teacher in self.teacher_schedule[day][h] for h in range(0,2))
                            last_two_hours_allocated = any(teacher in self.teacher_schedule[day][h] for h in range(7, 9))
                            lab_last2=any(h in range(7, 9) for _, h in lab_time_slots)
                            lab_first2=any(h in range(0,2) for _, h in lab_time_slots)

                            # If the teacher is assigned to the first two hours, do not assign to the last two hours and vice versa
                            if first_two_hours_allocated and lab_last2:
                                
                                continue  # Skip assigning teacher if there's a conflict with last two hours
                            if last_two_hours_allocated and lab_first2:
                                continue  # Skip assigning teacher if there's a conflict with first two hours

                            # Assign the teacher to this lab session if no conflicts
                            allocated_teachers.append(teacher)
                            teacher_hours[teacher] = current_hours + 2 # Update teacher's hours

                        # Stop once we have 5 teachers
                        if len(allocated_teachers) >= 5:
                            break

                # Ensure only unique teachers are assigned
                allocated_teachers = list(set(allocated_teachers))

                # If less than 5 teachers are allocated, notify
                if len(allocated_teachers) <= 4:
                    allocated_teachers=self.assign_teachers_to_unassigned_subjects(timetable,allocated_teachers ,class_name, subject_name, lab_time_slots)
                    

                    # print(f"Only {len(allocated_teachers)} teachers allocated for {class_name} - {subject_name} lab. Need more teachers.")

                # Assign teachers to the first 2 lab time slots
                self.lab_teachers[class_name][subject_name].append(allocated_teachers)
                for (day, hour) in lab_time_slots:
                    if hour <= 8:  # Ensure valid hour
                       
                        if allocated_teachers:  # Only proceed if we have allocated teachers
                            # Update the timetable with allocated teachers
                           
                            existing_entry = timetable[day][class_name][hour][0]
                            subject_name = existing_entry[0]  # Subject name remains the same
                            current_teachers = existing_entry[1]  # List of current teachers
                            lab_name = existing_entry[2]  # Lab name remains the same

                            # Append new teachers to the current list of teachers
                            updated_teachers = list(set(current_teachers + allocated_teachers))

                            # Update the timetable with the new teacher list, keeping the subject and lab unchanged
                            timetable[day][class_name][hour][0] = (subject_name, updated_teachers, lab_name)
                            # Block the teacher's schedule for the allocated time slots
                            for teacher in allocated_teachers:
                                self.teacher_schedule[day][hour].add(teacher)  # Block the teacher's time
                if len(allocated_teachers) <= 4:
                    print(f"Only {len(allocated_teachers)} teachers could be allocated for {class_name} - {subject_name} lab.")
        
    def assign_teachers_to_unassigned_subjects(self, timetable, allocated_teachers,class_name, subject_name, lab_time_slots):
            """
            Assigns teachers to subjects that did not meet the allocation criteria previously.
            Updates the timetable and blocks the teacher's schedule for all relevant days.

            Args:
            timetable (list): The timetable structure.
            class_name (str): The name of the class.
            subject_name (str): The name of the subject.
            lab_time_slots (list): The available time slots for the lab.
            """
            teacher_free_hours = self.get_teacher_free_hours(self.num_days, self.num_hours)
            teacher_hours = self.calculate_teacher_hours()

            for teacher, free_slots in teacher_free_hours.items():
                free_slots_set = set(free_slots)

                # Check teacher-specific hour limits
                if ((teacher.lower() == "dr sagaya aurelia p".lower() and teacher_hours.get(teacher, 0) >= 12) or 
                    (teacher.lower() == "dr manjunatha hiremath".lower() and teacher_hours.get(teacher, 0) >= 12)):
                    continue  # Skip if the teacher has reached their limit

                # Ensure the teacher has enough free slots for the lab
                lab_time_slots_set = set(lab_time_slots[:2])  # Assuming we want to fill the first two slots
                if lab_time_slots_set.issubset(free_slots_set):
                    current_hours = teacher_hours.get(teacher, 0)

                    # Ensure adding this lab won't exceed 18 hours for the teacher
                    if current_hours + 2 <= 18:
                        
                         # Skip assigning teacher if there's a conflict with first two hours

                        # Block the teacher's schedule for the allocated time slots
                        for (day, hour) in lab_time_slots:
                            first_two_hours_allocated = any(teacher in self.teacher_schedule[day][h] for h in range(0,2))
                            last_two_hours_allocated = any(teacher in self.teacher_schedule[day][h] for h in range(7, 9))
                            lab_last2=any(h in range(7, 9) for _, h in lab_time_slots)
                            lab_first2=any(h in range(0,2) for _, h in lab_time_slots)

                            # If the teacher is assigned to the first two hours, do not assign to the last two hours and vice versa
                            if first_two_hours_allocated and lab_last2:
                                
                                continue  # Skip assigning teacher if there's a conflict with last two hours
                            if last_two_hours_allocated and lab_first2:
                                continue 
                            
                            if hour <= 8:  # Ensure valid hour
                                # Only add teacher to timetable if they are not already scheduled
                                if teacher not in self.teacher_schedule[day][hour]:
                                    # Allocate teacher to timetable
                                    # timetable[day][class_name][hour].append(("", [teacher]))  # Assign teacher to the timetable
                                    self.teacher_schedule[day][hour].add(teacher)  # Block the teacher's time
                                    allocated_teachers.append(teacher)  # Track allocated teachers

                                    # Update teacher's hours
                                    teacher_hours[teacher] = current_hours + 2

                                    # Stop once we have 5 teachers
                                    if len(allocated_teachers) >= 5:
                                        break

                # Stop if 5 teachers have already been allocated
                if len(allocated_teachers) >= 5:
                    break

            # If still less than 5 teachers are allocated, notify
            # if len(allocated_teachers) < 5:
            #     self.assign_teachers_to_unassigned_subjects(timetable, allocated_teachers,class_name, subject_name, lab_time_slots)

                # print(f"Only {len(allocated_teachers)} teachers could be allocated for {class_name} - {subject_name} lab.")
            
            return allocated_teachers


    
    def is_teacher_allocated_in_first_two_hours(self, teacher_names, day, hour):
        """
        Check if the given hour is the last or second last hour of the day
        and whether any of the specified teachers are allocated in the first or second hour.

        Args:
        teacher_names (list): A list of teacher names to check.
        day (int): The day index (0-based).
        hour (int): The hour index (0-based).

        Returns:
        bool: False if the hour is last or second last and any teacher is allocated in the first two hours, else True.
        """
        # Check if the hour is last or second last
        if hour >= (self.num_hours - 2):  # Assuming self.num_hours is defined
            # Check if any teacher in the list is allocated in the first or second hour
            first_hour_allocated = any(teacher in self.teacher_schedule[day][0] for teacher in teacher_names)  # Check first hour (index 0)
            second_hour_allocated = any(teacher in self.teacher_schedule[day][1] for teacher in teacher_names)  # Check second hour (index 1)

            if first_hour_allocated or second_hour_allocated:
                return False  # At least one teacher is allocated in the first two hours

        return True  # No teachers are allocated in the first two hours, or hour is not last/second last


    
    
    def cross_verify_timetable_with_courses(self, timetable):
        """
        Traverses the timetable and creates a dictionary with (class, subject) as key
        and the number of hours taught as the value. Prints classes with irregularities 
        in assigned hours.

        Args:
        timetable (list): The timetable structure where timetable[day][class_name][hour] contains subjects.
        
        Returns:
        None: Prints classes and subjects with irregular assigned hours.
        """
        hours_taught = {}  # Dictionary to store hours taught for each (class, subject)
        
        # Traverse the timetable to count the hours each subject is taught
        for day in range(self.num_days):  # Assuming `self.num_days` is defined
            for class_name in timetable[day]:  # Traverse each class in the timetable for that day
                for hour in range(self.num_hours):  # Traverse each hour
                    for subject_info in timetable[day][class_name][hour]:
                        subject_name = subject_info[0]  # Assuming the subject name is the first element
                        
                        # Create key (class_name, subject_name)
                        key = (class_name, subject_name)
                        
                        # Count the number of hours taught for this subject
                        if key in hours_taught:
                            hours_taught[key] += 1
                        else:
                            hours_taught[key] = 1

        # Compare the hours taught with the courses dictionary and print irregularities
        for class_name, class_data in self.courses.items():
            for category, subjects in class_data.items():
                for subject_name, subject_info in subjects.items():
                    required_hours = subject_info["total_hours_per_week"]
                    
                    # Key to lookup in hours_taught
                    key = (class_name, subject_name)
                    
                    # Check if the subject is in hours_taught and compare with required hours
                    if key in hours_taught:
                        if hours_taught[key] != required_hours:
                            print(f"Class {class_name} - Subject '{subject_name}' has {hours_taught[key]} assigned hours, "
                                f"but requires {required_hours} hours.")
                    else:
                        print(f"Class {class_name} - Subject '{subject_name}' is not assigned in the timetable.")


  

    def get_teacher_free_hours(self, num_days=6, num_hours=8):
        """
        Retrieves all free hours for each teacher from the teacher's schedule, sorted by the number of free hours.

        Args:
        num_days (int): The number of days in the schedule.
        num_hours (int): The number of hours in the schedule for each day.

        Returns:
        dict: A dictionary where the key is the teacher's name and the value is a list of tuples (day, hour) representing their free hours, sorted by the number of free hours.
        """
        # Dictionary to store free hours for each teacher
        teacher_free_hours = {}
        # Union of all known teachers from the schedule
        all_teachers = set(teacher for schedule in self.teacher_schedule.values() 
                                for hour_teachers in schedule.values() 
                                for teacher in hour_teachers)
        # all_teachers = set(short_teachers.keys())
        # Loop over all possible days and hours
        for day in range(num_days):
            for hour in range(num_hours):
                # Get the set of busy teachers for the current day/hour
                busy_teachers = self.teacher_schedule.get(day, {}).get(hour, set())

                

                # Loop through all teachers
                for teacher in all_teachers:
                    if teacher.lower().startswith("dr"):  # Only consider teachers starting with "Dr"
                        # Initialize the teacher's free hours list if not already done
                        if teacher not in teacher_free_hours:
                            teacher_free_hours[teacher] = []

                        # If the teacher is not busy at this time, mark it as free
                        if teacher not in busy_teachers:
                            teacher_free_hours[teacher].append((day, hour))

      
        sorted_teacher_free_hours = dict(
            sorted(teacher_free_hours.items(), key=lambda item: len(item[1]), reverse=True)
        )

        return sorted_teacher_free_hours


    
    
    
    def get_lab_hours(self,timetable):
        """
        Retrieves all lab hours for each (class_name, subject_name) in the timetable.

        Args:
        timetable (list): A list of lists where timetable[day][class_name][hour] contains subjects for each hour in each class.

        Returns:
        dict: A dictionary where the key is a tuple (class_name, subject_name) and the value is a list of tuples (day, hour) representing the lab timings.
        """
        lab_hours = {}

        for day in range(len(timetable)):  # Iterate over each day
            for class_name in timetable[day]:  # Iterate over each class
                for hour in range(len(timetable[day][class_name])):  # Iterate over each hour
                    subjects = timetable[day][class_name][hour]

                    for subject in subjects:
                        # Check if it's a lab (lab subjects are usually indicated by having 3 elements in the tuple)
                        if len(subject) == 3:  # Subject tuple (subject_name, teachers, lab_venue)
                            subject_name = subject[0]  # First element is the subject name
                            lab_venue = subject[2]     # Third element is the lab venue

                            # Create the key (class_name, subject_name)
                            key = (class_name, subject_name)

                            # Store the day and hour as lab timing
                            if key in lab_hours:
                                lab_hours[key].append((day, hour))
                            else:
                                lab_hours[key] = [(day, hour)]

        return lab_hours

    
    def print_teacher_hr(self,teacherhr):
        i=1
        for x,y in teacherhr.items():
            print(f"{i} {x} :{y} ")
            i=i+1
           
            
        
    
    def calculate_teacher_hours(self):
        """
        Calculates the total number of teaching hours for each teacher based on their schedule.

        Returns:
        dict: A dictionary where the key is the teacher's name and the value is the number of hours they are teaching.
        """
        teacher_hours = {}

        # Traverse through each day and each hour in the teacher's schedule
        for day in range(self.num_days):  # Assuming num_days is defined
            for hour in range(self.num_hours):  # Assuming num_hours is defined
                # Check which teachers are scheduled for this hour
                for teacher in self.teacher_schedule[day][hour]:
                    teacher = teacher.strip()  # Clean any extra whitespace

                    # Increment the count of teaching hours for this teacher
                    if teacher.lower().startswith("dr"):  # Check for title if necessary
                        if teacher in teacher_hours:
                            teacher_hours[teacher] += 1
                        else:
                            teacher_hours[teacher] = 1
        sorted_dict = dict(sorted(teacher_hours.items(), key=lambda item: item[1], reverse=True))
        return sorted_dict


        

    def initialize_timetable(self):
        return {
            day: {class_name: {hour: [] for hour in range(self.num_hours)} for class_name in self.classes}
            for day in range(self.num_days)
        }

    def initialize_tracking_structures(self):
        allocated_subjects = {day: {class_name: set() for class_name in self.classes} for day in range(self.num_days)}
        theory_hours_assigned = {day: {class_name: {} for class_name in self.classes} for day in range(self.num_days)}
        lab_assigned_per_day = {day: {class_name: False for class_name in self.classes} for day in range(self.num_days)}
        teacher_schedule = {day: {hour: set() for hour in range(self.num_hours)} for day in range(self.num_days)}
        return allocated_subjects, theory_hours_assigned, lab_assigned_per_day, teacher_schedule
    
    def pre_assign_subjects(self, timetable):
        """Pre-assign subjects to the timetable."""
        
        for class_name in self.classes:
           
            self.assign_hed(timetable, class_name)
            self.assign_mdc(timetable, class_name)
            self.assign_act(timetable,class_name)
            self.block_hours(timetable, class_name)
    def pre_assign_extra_subjects(self,timetable,d):
        for class_name in self.classes:
            self.commerce(timetable,class_name)
            self.acc(timetable,class_name)
            self.assign_lunch(timetable, class_name)
            self.assign_language(timetable,class_name)
            self.assign_lib(timetable,class_name,d)
            
    def assign_lib(self,timetable,class_name,d):
        # print(d)
        # if "LIBRARY" in courses[class_name]:
        if  not timetable[1][class_name][5]:
            timetable[d][class_name][5].append(("Library",["lib"]))
            
    def acc(self,timetable,class_name):
        if class_name in ["3BCA A","3BCA B"]:
            timetable[1][class_name][5].append(("Accounting",["acc"]))
            timetable[4][class_name][5].append(("Accounting",["acc"]))
            timetable[4][class_name][4].append(("Accounting",["acc"]))
            
    def commerce(self, timetable, class_name):
        if class_name in ["BCOM-I", "BCOM-II", "BCOM-III"]:
            l = None 
            if class_name == "BCOM-I":
                l = list1
            elif class_name == "BCOM-II":
                l = list2
            elif class_name == "BCOM-III":
                l = list3

            for d in range(4, 6):
                for hr in range(2, 4):
                    if d == 4:
                        self.teacher_schedule[d][hr-2].add(l[0])  # Block the teacher's schedule                 
                        timetable[d][class_name][hr-2].append((class_name, [l[0]]))
                    elif d == 5:
                        for x in l:
                            self.teacher_schedule[d][hr].add(x)  # Block the teacher's schedule
                        timetable[d][class_name][hr].append((class_name, l))
                            
        
        
        
           
    def assign_act(self, timetable, class_name):
        """Pre-assign 'activity' to the specific periods on Wednesday for different classes."""
        
        # if 'act' in self.courses[class_name]:
        timetable[2][class_name][5].append(("activity", [""]))
  
            
    def assign_language(self, timetable, class_name):
        """Pre-assign 'language' to specific periods on Thursday and Friday for all classes."""
        
        if 'language' in self.courses[class_name]:
            if class_name in n_m:
                # Thursday (5th period)
                timetable[4][class_name][4].append(("language", []))
                # Friday (4th period)
                timetable[4][class_name][3].append(("language", [])) 
            elif class_name in m:
                if class_name in ["3CS","3CM"]:
                    
                     # wed (5th period)
                    timetable[2][class_name][4].append(("language", []))
                    # Thursday (6th period)
                    timetable[4][class_name][5].append(("language", [])) 
                     # Tuesday (4th period)
                    
                else:
                    timetable[1][class_name][3].append(("language", []))
                    # Thursday (4th period)
                    timetable[3][class_name][3].append(("language", [])) 
               

    def assign_hed(self, timetable, class_name):
        """Pre-assign 'HED' to the 6th hour for all classes."""
        if 'HED' in self.courses[class_name]:
            timetable[0][class_name][5].append(("HED", []))  # 5th hour

    def assign_mdc(self, timetable, class_name):
        """Pre-assign 'MDC' to specific hours if the class has it."""
        if 'MDC' in self.courses[class_name]:  # Check if "MDC" exists for this class
            for hour in range(3, 5):  # 3rd and 4th hours on day 1
                timetable[1][class_name][hour].append(("MDC", []))
            # Pre-assign "MDC" to the 2nd hour on day 4
            timetable[3][class_name][3].append(("MDC", []))  # 2nd hour on day 4

    def assign_lunch(self, timetable, class_name):
        """Assign lunch based on the 'NOT MORNING' key."""
        for day in range(self.num_days):
            if day == 5:  # Skip Saturday
                continue
            if class_name in n_m:
                timetable[day][class_name][6].append(("LUNCH", []))  # 7th hour for lunch
            else:
                timetable[day][class_name][2].append(("BREAK", []))  # 3rd hour for lunch

    def block_hours(self, timetable, class_name):
        """
        Blocks specific hours in a timetable based on the 'NOT MORNING' constraint.

        Parameters:
        timetable (dict): A dictionary representing the timetable, where each class name 
                        corresponds to a list of hours for each day.
        class_name (str): The name of the class for which the hours will be blocked.

        Behavior:
        - If the class is in the 'n_m' (NOT MORNING) list, the function will block the 
        first 3 hours of all days except day 5. Additionally, it will block specific 
        hours (0, 5, 6, 7, 8) for day 5.
        - For classes that are not restricted by 'NOT MORNING' (i.e., not in 'n_m'), it 
        blocks hours after the 5th hour across all days. An exception is made for classes 
        "5BCA A" and "5BCA B" where hour 5 on day 5 is not blocked.

        Conditions:
        - The class must be in the 'NOT MORNING' (n_m) list to block the first few hours.
        - Classes that do not meet this condition but do not have 'NOT MORNING' in their
        metadata ('m') will have later hours blocked instead.
        """
        if class_name not in ["5BCA A", "5BCA B"]:
                timetable[5][class_name][6].append(("BLOCKED", []))
        
        for hour in [0,1,7, 8]:
            
                timetable[5][class_name][hour].append(("BLOCKED", []))
        
        if class_name in n_m:
            # Block first 3 hours for all days except day 5
            for day in range(self.num_days):
                if day != 5:
                    for hour in range(3):
                        timetable[day][class_name][hour].append(("BLOCKED", []))

           

        # Block all hours after the 5th for classes without 'NOT MORNING'
        elif "NOT MORNING" not in m:
            # Always block hour 0 on day 5
            # timetable[5][class_name][0].append(("BLOCKED", []))

            # Block hour 5 on day 5 unless the class is "5BCA A" or "5BCA B"
            if class_name not in ["5BCA A", "5BCA B"]:
                timetable[5][class_name][6].append(("BLOCKED", []))

            # Block hours from the 6th hour onward for all days
            for hour in range(6, len(timetable[0][class_name])):
                for day in range(self.num_days):
                    if day!=5:
                        timetable[day][class_name][hour].append(("BLOCKED", []))
        





    def assign_lca_lab_hours(self, timetable, allocated_subjects, lab_assigned_per_day):
        """
        Assigns lab hours to the timetable based on the number of hours mentioned for each subject.

        Args:
        timetable (list): The timetable structure.
        allocated_subjects (set): The set of subjects already allocated in the timetable.
        lab_assigned_per_day (list): A boolean list that tracks whether a lab has been assigned for each class per day.
        """
        for class_name in self.classes:
            if 'LCA' in self.courses[class_name]:  # Check if LCA exists for this class
                lca_subjects = self.courses[class_name]['LCA']

                for subject_name, subject_info in lca_subjects.items():
                    lab_hours = subject_info["lab_hours"]  # Total required lab hours
                    teachers_for_subject = subject_info["teacher_incharge"]
                    assigned_lab_hours = 0

                    for day in range(self.num_days):
                        if assigned_lab_hours >= lab_hours:
                            break  # Stop if we've assigned all required lab hours

                        # Ensure that the lab hasn't been assigned already and the first two hours are free
                        if not lab_assigned_per_day[day][class_name]:
                            if len(timetable[day][class_name][0]) == 0 and len(timetable[day][class_name][1]) == 0:
                                # Check if teachers are available
                                if (self.is_teacher_allocated_in_first_two_hours(teachers_for_subject, day, 0) and
                                    all(teacher not in self.teacher_schedule[day][0] for teacher in teachers_for_subject)):

                                    lab_venue = self.get_available_lab(day, 0)  # Get an available lab venue
                                    if lab_venue and self.is_lab_available(lab_venue, day, 0):  # Check if the lab is available
                                        # Assign the lab to the timetable and block the teacher's schedule
                                        self.assign_lab(timetable, day, class_name, 0, subject_name, teachers_for_subject, lab_venue, allocated_subjects, self.teacher_schedule)
                                        lab_assigned_per_day[day][class_name] = True
                                        assigned_lab_hours += 2  # Increment assigned lab hours by 2
                                        break  # Move to the next subject once lab hours are assigned

    def assign_lca_hours(self, timetable, allocated_subjects):
        """
        Assigns LCA (normal) hours to the timetable based on the number of hours mentioned for each subject.

        Args:
        timetable (list): The timetable structure.
        allocated_subjects (set): The set of subjects already allocated in the timetable.
        """
        for class_name in self.classes:
            if 'LCA' in self.courses[class_name]:  # Check if LCA exists for this class
                lca_subjects = self.courses[class_name]['LCA']

                for subject_name, subject_info in lca_subjects.items():
                    required_hours = subject_info["normal_hours"]  # Required LCA hours
                    assigned_lca_hours = 0

                    for day in range(self.num_days):
                        if assigned_lca_hours >= required_hours:
                            break  # Stop if all required LCA hours have been assigned

                        # Check if we can assign two consecutive hours in the first two hours of the day
                        for hour in range(2):  # Only check hours 0 and 1
                            teachers_for_lca = subject_info["teacher_incharge"]
                            can_assign = (len(timetable[day][class_name][hour]) == 0 and 
                                        len(timetable[day][class_name][hour + 1]) == 0 and
                                        all(teacher not in self.teacher_schedule[day][hour] and 
                                            teacher not in self.teacher_schedule[day][hour + 1]
                                            for teacher in teachers_for_lca))

                            if can_assign:
                                if assigned_lca_hours + 1 <= required_hours:
                                    # Assign the first hour
                                    timetable[day][class_name][hour].append((subject_name, teachers_for_lca))
                                    for teacher in teachers_for_lca:
                                        self.teacher_schedule[day][hour].add(teacher)  # Block the teacher's schedule
                                    assigned_lca_hours += 1

                                if assigned_lca_hours < required_hours and assigned_lca_hours + 1 <= required_hours:
                                    # Assign the second hour
                                    timetable[day][class_name][hour + 1].append((subject_name, teachers_for_lca))
                                    for teacher in teachers_for_lca:
                                        self.teacher_schedule[day][hour + 1].add(teacher)  # Block the teacher's schedule
                                    assigned_lca_hours += 1

                                break  # Move to the next day after assigning hours

                        # Stop if all LCA hours are assigned
                        if assigned_lca_hours >= required_hours:
                            break


    def assign_elective_subject(self, timetable, day, class_name, hour, subject_name, teachers_for_subject, teacher_schedule):
        # Check if the subject is already assigned for this hour
        if any(subject[0] == subject_name for subject in timetable[day][class_name][hour]):
            return False

        # Ensure teachers_for_subject is a list
        available_teachers = teachers_for_subject if isinstance(teachers_for_subject, list) else [teachers_for_subject]

        # Select a random teacher from available teachers
        if available_teachers:
            teacher = random.choice(available_teachers)
            timetable[day][class_name][hour].append((subject_name, [teacher]))
            self.teacher_schedule[day][hour].add(teacher)
            return True
        return False

    def assign_elective_hours(self, timetable, allocated_subjects, teacher_schedule):
        theory_hours_assigned = {day: {class_name: {} for class_name in self.courses.keys()} for day in range(self.num_days)}

        for class_name, class_data in self.courses.items():
            if class_name in [ele[0]]:
                for category, subjects in class_data.items():
                    if category in ["ELECTIVE-I", "ELECTIVE-II"]:
                        elective_subjects = list(subjects.keys())
                        if len(elective_subjects) < 2:
                            continue  # Skip if less than 2 subjects

                        subject1_name = elective_subjects[0]
                        subject1_info = subjects[subject1_name]
                        subject2_name = elective_subjects[1]
                        subject2_info = subjects[subject2_name]

                        normal_hours1 = subject1_info["normal_hours"]
                        normal_hours2 = subject2_info["normal_hours"]
                        teachers_for_subject1 = subject1_info["teacher_incharge"]
                        teachers_for_subject2 = subject2_info["teacher_incharge"]

                        assigned_theory_hours1 = 0
                        assigned_theory_hours2 = 0

                        while assigned_theory_hours1 < normal_hours1 or assigned_theory_hours2 < normal_hours2:
                            for day in range(self.num_days):
                                total_hours1_A = theory_hours_assigned[day][ele[0]].get(subject1_name, 0)
                                total_hours1_B = theory_hours_assigned[day][ele[1]].get(subject1_name, 0)
                                total_hours2_A = theory_hours_assigned[day][ele[0]].get(subject2_name, 0)
                                total_hours2_B = theory_hours_assigned[day][ele[1]].get(subject2_name, 0)

                                # Check availability for both subjects
                                if (total_hours1_A < 2 and total_hours1_B < 2 and 
                                    total_hours2_A < 2 and total_hours2_B < 2 and 
                                    subject1_name not in allocated_subjects[day][ele[0]] and 
                                    subject1_name not in allocated_subjects[day][ele[1]] and 
                                    subject2_name not in allocated_subjects[day][ele[0]] and 
                                    subject2_name not in allocated_subjects[day][ele[1]]):
                                    
                                    for hour in range(self.num_hours):
                                        if (len(timetable[day][ele[0]][hour]) == 0 and 
                                            len(timetable[day][ele[1]][hour]) == 0):
                                            
                                            # Check if teachers are free for both subjects
                                            teachers_available = (all(teacher not in self.teacher_schedule[day][hour] for teacher in teachers_for_subject1) and 
                                                                all(teacher not in self.teacher_schedule[day][hour] for teacher in teachers_for_subject2))
                                            
                                            p=self.is_teacher_allocated_in_first_two_hours(teachers_for_subject1,day,hour)
                                            p2=self.is_teacher_allocated_in_first_two_hours(teachers_for_subject2,day,hour)

                                            if teachers_available and p and p2:
                                                # Assign subject 1 if hours are available
                                                if assigned_theory_hours1 < normal_hours1:
                                                    if self.assign_elective_subject(timetable, day, ele[0], hour, subject1_name, teachers_for_subject1, teacher_schedule) and \
                                                    self.assign_elective_subject(timetable, day, ele[1], hour, subject1_name, teachers_for_subject1, teacher_schedule):
                                                        
                                                        for teacher in teachers_for_subject1:
                                                            self.teacher_schedule[day][hour].add(teacher)

                                                        theory_hours_assigned[day][ele[0]][subject1_name] = total_hours1_A + 1
                                                        allocated_subjects[day][ele[0]].add(subject1_name)
                                                        theory_hours_assigned[day][ele[1]][subject1_name] = total_hours1_B + 1
                                                        allocated_subjects[day][ele[1]].add(subject1_name)
                                                        assigned_theory_hours1 += 1

                                                # Assign subject 2 if hours are available
                                                if assigned_theory_hours2 < normal_hours2:
                                                    if self.assign_elective_subject(timetable, day, ele[0], hour, subject2_name, teachers_for_subject2, teacher_schedule) and \
                                                    self.assign_elective_subject(timetable, day, ele[1], hour, subject2_name, teachers_for_subject2, teacher_schedule):
                                                        
                                                        for teacher in teachers_for_subject2:
                                                            self.teacher_schedule[day][hour].add(teacher)

                                                        theory_hours_assigned[day][ele[0]][subject2_name] = total_hours2_A + 1
                                                        allocated_subjects[day][ele[0]].add(subject2_name)
                                                        theory_hours_assigned[day][ele[1]][subject2_name] = total_hours2_B + 1
                                                        allocated_subjects[day][ele[1]].add(subject2_name)
                                                        assigned_theory_hours2 += 1

                                                break  # Exit for loop if assignments are made

                                # Check if both subjects have been assigned the required hours
                                if assigned_theory_hours1 >= normal_hours1 and assigned_theory_hours2 >= normal_hours2:
                                    break
    def assign_elective_lab(self, timetable, allocated_subjects, lab_assigned_per_day, teacher_schedule):
        for class_name, class_data in self.courses.items():
            if class_name in ['5BCA A']:
                for category, subjects in class_data.items():
                    if category == "ELECTIVE-I":
                        assigned_lab_hours_ga = 0
                        assigned_lab_hours_bi = 0
                        # Prepare subject names
                        subjects_to_assign = [i for i in courses[ele[0]]["ELECTIVE-I"].keys()]

                        ga_lab_hours_needed = subjects[subjects_to_assign[0]]["lab_hours"]
                        bi_lab_hours_needed = subjects[subjects_to_assign[1]]["lab_hours"]

                        for day in range(self.num_days):
                            if (assigned_lab_hours_ga >= ga_lab_hours_needed and
                                    assigned_lab_hours_bi >= bi_lab_hours_needed):
                                break

                            # Check if both subjects can be assigned on the same day
                            if all(subject_name not in allocated_subjects[day][class_name] for subject_name in subjects_to_assign) and not lab_assigned_per_day[day][class_name]:
                                for hour in range(self.num_hours - 1):  # Ensure space for two hours

                                    # Check if both labs can be assigned in the same time slot
                                    if assigned_lab_hours_ga < ga_lab_hours_needed and assigned_lab_hours_bi < bi_lab_hours_needed:
                                        
                                        # Check if the time slots are free in the timetable for both subjects
                                        if len(timetable[day][class_name][hour]) == 0 and len(timetable[day][class_name][hour + 1]) == 0 and self.labs_assigned_per_hour[day][hour]==0:
                                            
                                            # Check teacher availability for both subjects
                                            teachers_for_ga = subjects[subjects_to_assign[0]]["teacher_incharge"]
                                            teachers_for_bi = subjects[subjects_to_assign[1]]["teacher_incharge"]

                                            teachers_available = (all(teacher not in self.teacher_schedule[day][hour] for teacher in teachers_for_ga) and 
                                                                all(teacher not in self.teacher_schedule[day][hour] for teacher in teachers_for_bi))

                                            p_ga = self.is_teacher_allocated_in_first_two_hours(teachers_for_ga, day, hour)
                                            p_bi = self.is_teacher_allocated_in_first_two_hours(teachers_for_bi, day, hour)

                                            if teachers_available and p_ga and p_bi:
                                                # Try to assign labs for both subjects
                                                labs = self.get_two_different_available_labs(day, hour)
                                                lab_venue_ga = None
                                                lab_venue_bi = None
                                                if labs:
                                                    lab_venue_ga = labs[0]
                                                    lab_venue_bi = labs[1]
                                                if lab_venue_ga and self.is_lab_available(lab_venue_ga, day, hour) and \
                                                lab_venue_bi and self.is_lab_available(lab_venue_bi, day, hour):
                                                    
                                                    # Assign lab for 5BCA A - GA
                                                    self.assign_lab(timetable, day, class_name, hour, subjects_to_assign[0],
                                                                    teachers_for_ga, lab_venue_ga, allocated_subjects, teacher_schedule)
                                                    lab_assigned_per_day[day][class_name] = True
                                                    for teacher in teachers_for_ga:
                                                        self.teacher_schedule[day][hour].add(teacher)

                                                    # Assign lab for 5BCA B - GA
                                                    self.assign_lab(timetable, day, "5BCA B", hour, subjects_to_assign[0],
                                                                    teachers_for_ga, lab_venue_ga, allocated_subjects, teacher_schedule)
                                                    lab_assigned_per_day[day]["5BCA B"] = True
                                                    for teacher in teachers_for_ga:
                                                        self.teacher_schedule[day][hour].add(teacher)

                                                    # Assign lab for 5BCA A - BI (in a different venue if available)
                                                    self.assign_lab(timetable, day, class_name, hour , subjects_to_assign[1],
                                                                    teachers_for_bi, lab_venue_bi, allocated_subjects, teacher_schedule)
                                                    lab_assigned_per_day[day][class_name] = True
                                                    for teacher in teachers_for_bi:
                                                        self.teacher_schedule[day][hour + 1].add(teacher)

                                                    # Assign lab for 5BCA B - BI
                                                    self.assign_lab(timetable, day, "5BCA B", hour , subjects_to_assign[1],
                                                                    teachers_for_bi, lab_venue_bi, allocated_subjects, teacher_schedule)
                                                    lab_assigned_per_day[day]["5BCA B"] = True
                                                    for teacher in teachers_for_bi:
                                                        self.teacher_schedule[day][hour + 1].add(teacher)

                                                    # Update assigned lab hours
                                                    assigned_lab_hours_ga += 2  # Two hours for GA
                                                    assigned_lab_hours_bi += 2  # Two hours for BI
                                                    self.labs_assigned_per_hour[day][hour]+=2
                                                    self.labs_assigned_per_hour[day][hour+1]+=2
                                                    
                                                    

                                                    # print(f"Assigned GA Lab Hours: {assigned_lab_hours_ga} / {ga_lab_hours_needed}")
                                                    # print(f"Assigned BI Lab Hours: {assigned_lab_hours_bi} / {bi_lab_hours_needed}")

                                                    # # Exit the loop after successful assignment
                                                    break



    
    def assign_elective_hours_single_section(self, timetable, allocated_subjects, teacher_schedule, class_name):
        theory_hours_assigned = {day: {class_name: {} for day in range(self.num_days)} for day in range(self.num_days)}

        class_data = self.courses[class_name]
        for category, subjects in class_data.items():
            if category == "ELECTIVE-I":
                elective_subjects = list(subjects.keys())
                if len(elective_subjects) < 2:
                    continue  # Skip if less than 2 subjects

                subject1_name, subject2_name = elective_subjects[0], elective_subjects[1]
                subject1_info, subject2_info = subjects[subject1_name], subjects[subject2_name]

                normal_hours1, normal_hours2 = subject1_info["normal_hours"], subject2_info["normal_hours"]
                teachers_for_subject1, teachers_for_subject2 = subject1_info["teacher_incharge"], subject2_info["teacher_incharge"]

                assigned_theory_hours1, assigned_theory_hours2 = 0, 0

                while assigned_theory_hours1 < normal_hours1 or assigned_theory_hours2 < normal_hours2:
                    for day in range(self.num_days):
                        total_hours1 = theory_hours_assigned[day][class_name].get(subject1_name, 0)
                        total_hours2 = theory_hours_assigned[day][class_name].get(subject2_name, 0)

                        if (total_hours1 < 2 and total_hours2 < 2 and
                            subject1_name not in allocated_subjects[day][class_name] and
                            subject2_name not in allocated_subjects[day][class_name]):

                            for hour in range(self.num_hours):
                                if len(timetable[day][class_name][hour]) == 0:  # If the hour is free
                                    # Check if the teacher is free for both subjects
                                    p=self.is_teacher_allocated_in_first_two_hours(teachers_for_subject1,day,hour)
                                    p2=self.is_teacher_allocated_in_first_two_hours(teachers_for_subject2,day,hour)
                                    ava=(all(teacher not in self.teacher_schedule[day][hour] for teacher in teachers_for_subject1) and
                                        all(teacher not in self.teacher_schedule[day][hour] for teacher in teachers_for_subject2))
                                    if ava and p and p2 :
                                        
                                        # Assign subject 1
                                        if assigned_theory_hours1 < normal_hours1:
                                            self.assign_elective_subject(timetable, day, class_name, hour, subject1_name, teachers_for_subject1, teacher_schedule)
                                            for teacher in teachers_for_subject1:
                                                self.teacher_schedule[day][hour].add(teacher)
                                            
                                            theory_hours_assigned[day][class_name][subject1_name] = total_hours1 + 1
                                            allocated_subjects[day][class_name].add(subject1_name)
                                            assigned_theory_hours1 += 1

                                        # Assign subject 2
                                        if assigned_theory_hours2 < normal_hours2:
                                            self.assign_elective_subject(timetable, day, class_name, hour, subject2_name, teachers_for_subject2, teacher_schedule)
                                            for teacher in teachers_for_subject2:
                                                self.teacher_schedule[day][hour].add(teacher)
                                            theory_hours_assigned[day][class_name][subject2_name] = total_hours2 + 1
                                            allocated_subjects[day][class_name].add(subject2_name)
                                            assigned_theory_hours2 += 1

                                        break  # Exit loop to re-evaluate assigned hours

                        if assigned_theory_hours1 >= normal_hours1 and assigned_theory_hours2 >= normal_hours2:
                            break

    def assign_lab_hours_multiple_electives(self, timetable, allocated_subjects, lab_assigned_per_day, teacher_schedule, class_name):
        for category, subjects in self.courses[class_name].items():
            if category == "ELECTIVE-I":
                elective_subjects = list(subjects.keys())
                if len(elective_subjects) < 2:
                    continue  # Skip if less than 2 subjects

                # Get the first two elective subjects
                subject1_name = elective_subjects[0]
                subject2_name = elective_subjects[1]
                subject1_info = subjects[subject1_name]
                subject2_info = subjects[subject2_name]

                lab_hours1 = subject1_info["lab_hours"]
                lab_hours2 = subject2_info["lab_hours"]
                teachers_for_subject1 = subject1_info["teacher_incharge"]
                teachers_for_subject2 = subject2_info["teacher_incharge"]

                assigned_lab_hours1 = 0
                assigned_lab_hours2 = 0

                for day in range(self.num_days):
                    if assigned_lab_hours1 >= lab_hours1 and assigned_lab_hours2 >= lab_hours2:
                        break

                    # Check if lab hours can be assigned for both subjects
                    if (subject1_name not in allocated_subjects[day][class_name] and 
                        subject2_name not in allocated_subjects[day][class_name] and
                        not lab_assigned_per_day[day][class_name]):
                        
                        for hour in range(self.num_hours - 1):  # Ensure space for two hours
                            # Check if both hours are free
                            if (len(timetable[day][class_name][hour]) == 0 and 
                                len(timetable[day][class_name][hour + 1]) == 0):
                                
                                ava= (all(teacher not in self.teacher_schedule[day][hour] for teacher in teachers_for_subject1) and
                                    all(teacher not in self.teacher_schedule[day][hour] for teacher in teachers_for_subject2))
                                    
                                p=self.is_teacher_allocated_in_first_two_hours(teachers_for_subject1,day,hour)
                                p2=self.is_teacher_allocated_in_first_two_hours(teachers_for_subject2,day,hour)

                                # Check if the teachers are free for both subjects
                                if   ava and p and p2 and self.labs_assigned_per_hour[day][hour]<=1 :  
                                    x=self.get_two_different_available_labs(day, hour)  # Assuming same lab for both subjects
                                    lab_venue1=x[0]
                                    lab_venue2=x[1]
                                    
                                    if lab_venue1 and self.is_lab_available(lab_venue1, day, hour):
                                        # Assign lab for both subjects in the same hour
                                        self.assign_lab(timetable, day, class_name, hour, subject1_name, teachers_for_subject1, lab_venue1, allocated_subjects, teacher_schedule)
                                        for teacher in teachers_for_subject1:
                                            self.teacher_schedule[day][hour].add(teacher)
                                    if lab_venue2 and self.is_lab_available(lab_venue2, day, hour):
                                        self.assign_lab(timetable, day, class_name, hour, subject2_name, teachers_for_subject2, lab_venue2, allocated_subjects, teacher_schedule)
                                        for teacher in teachers_for_subject2:
                                            self.teacher_schedule[day][hour].add(teacher)
                                        
                                        lab_assigned_per_day[day][class_name] = True
                                        assigned_lab_hours1 += 2  # Increment for subject 1
                                        assigned_lab_hours2 += 2  # Increment for subject 2
                                        self.labs_assigned_per_hour[day][hour]+=1
                                        break  # Exit the loop to re-evaluate lab assignments

                    if assigned_lab_hours1 >= lab_hours1 and assigned_lab_hours2 >= lab_hours2:
                        break  # Exit if all lab hours have been assigned
    def allocate_lab_alternatively(self, timetable, allocated_subjects, lab_assigned_per_day, teachers_for_subject, lab_hours, class_name, assigned_lab_hours, subject_name):  
        # Track assigned lab hours
        for day in range(0, self.num_days):        
            for hour in range(0, self.num_hours - 1):  # Ensure space for two hours
                if len(timetable[day][class_name][hour]) == 0 and len(timetable[day][class_name][hour + 1]) == 0:
                    # Check for alternative conditions
                    lab_venue = self.get_available_lab(day, hour)
                    if lab_venue and self.is_lab_available(lab_venue, day, hour):
                        # Assign lab since primary conditions weren't met
                        self.assign_lab(timetable, day, class_name, hour, subject_name, teachers_for_subject, lab_venue, allocated_subjects, self.teacher_schedule)
                        lab_assigned_per_day[day][class_name] = True
                        assigned_lab_hours += 2  # Increment by the number of lab hours assigned
                        if assigned_lab_hours >= lab_hours:
                            break  # Exit loop after successful allocation
            # If lab hours limit is reached, exit the outer loop as well
            if assigned_lab_hours >= lab_hours:
                break

        return assigned_lab_hours
  # Return the number of lab hours allocated


    def assign_lab_hours(self, timetable, allocated_subjects, lab_assigned_per_day, teacher_schedule):
        for class_name, class_data in self.courses.items():
            for category, subjects in class_data.items():
                if category not in ["LCA", "ELECTIVE-I", "HED", "MDC", "Elective-II", "act","LIBRARY"]:
                    for subject_name, subject_info in subjects.items():
                        
                        lab_hours = subject_info["lab_hours"]
                        teachers_for_subject = subject_info.get("teacher_incharge", [])
                        assigned_lab_hours = 0

                        for day in range(self.num_days):
                            if assigned_lab_hours >= lab_hours:
                                break
                            if subject_name not in allocated_subjects[day][class_name] and not lab_assigned_per_day[day][class_name]:
                                for hour in range(self.num_hours - 1):  # Ensure space for two hours
                                    if len(timetable[day][class_name][hour]) == 0 and len(timetable[day][class_name][hour + 1]) == 0:
                                        t = all(teacher not in self.teacher_schedule[day][hour] for teacher in teachers_for_subject)
                                        p = self.is_teacher_allocated_in_first_two_hours(teachers_for_subject, day, hour)
                                        if t and p and self.labs_assigned_per_hour[day][hour] <= 1:
                                            lab_venue = self.get_available_lab(day, hour)
                                            if lab_venue and self.is_lab_available(lab_venue, day, hour):
                                                self.assign_lab(timetable, day, class_name, hour, subject_name, teachers_for_subject, lab_venue, allocated_subjects, teacher_schedule)
                                                lab_assigned_per_day[day][class_name] = True
                                                assigned_lab_hours += 2
                                                self.labs_assigned_per_hour[day][hour] += 1
                                                break

                                # Attempt to allocate lab alternatively if not enough hours were assigned
                        if assigned_lab_hours <lab_hours:
                            print(f"Allocating lab alternatively for subject: {subject_name}")  # Debugging output
                            assigned_lab_hours = self.allocate_lab_alternatively(timetable, allocated_subjects, lab_assigned_per_day, teachers_for_subject, lab_hours-assigned_lab_hours, class_name, 0, subject_name)
                                        
                
    
    
    def assign_theory_hours(self, timetable, allocated_subjects, theory_hours_assigned, teacher_schedule):
        for class_name, class_data in self.courses.items():
            for category, subjects in class_data.items():
                if category not in ["LCA", "ELECTIVE-I",  "language","HED", "MDC","ELECTIVE-II","act"]:
                    for subject_name, subject_info in subjects.items():
                        normal_hours = subject_info["normal_hours"]
                        teachers_for_subject = subject_info["teacher_incharge"]

                        # Proceed only if normal_hours is greater than 0
                        if normal_hours > 0:
                            assigned_theory_hours = 0
                           
                            for day in range(self.num_days):
                                # Check if assigned theory hours have reached normal_hours limit
                                if assigned_theory_hours >= normal_hours:
                                    break

                                total_hours = theory_hours_assigned[day][class_name].get(subject_name, 0)

                                # Check if more hours can be assigned for this subject
                                if total_hours < 2 and subject_name not in allocated_subjects[day][class_name]:
                                    for hour in range(self.num_hours):
                                        if len(timetable[day][class_name][hour]) == 0:  # If the hour is free
                                            # Check if any teacher for the subject is available
                                            t=any(teacher not in self.teacher_schedule[day][hour] for teacher in teachers_for_subject)
                                            p=self.is_teacher_allocated_in_first_two_hours(teachers_for_subject,day,hour)
                                            if t and p:
                                                self.assign_subject(timetable, day, class_name, hour, subject_name, teachers_for_subject, category, teacher_schedule)
                                                for teacher in teachers_for_subject:
                                                    self.teacher_schedule[day][hour].add(teacher)
                                                    
                                                theory_hours_assigned[day][class_name][subject_name] = total_hours + 1
                                                allocated_subjects[day][class_name].add(subject_name)
                                                assigned_theory_hours += 1  # Increment assigned hours
                                                break  # Move to the next subject
                                    # print(class_name,subject_name,normal_hours,assigned_theory_hours)
                                    if assigned_theory_hours >= normal_hours:
                                        break  # Exit the loop once normal hours are assigned

                            # Check if the exact number of required theory hours was assigned
                            if assigned_theory_hours < normal_hours:
                                self.allocate_missing_hours_for_subject( class_name, subject_name, timetable, theory_hours_assigned, allocated_subjects, teacher_schedule,category)
                                # print(f"Warning: Could not assign the exact number of theory hours for {subject_name} in {class_name}. Assigned: {assigned_theory_hours}, Required: {normal_hours}")
                                
    
    def allocate_missing_hours_for_subject(self, class_name, subject_name, timetable, theory_hours_assigned, allocated_subjects, teacher_schedule,c="NORMAL"):
        # Access the relevant data for the specific class and subject
        subject_info = self.courses[class_name][c].get(subject_name)
        
        if not subject_info:
            print(f"Subject {subject_name} not found in class {class_name}.")
            return
        
        normal_hours = subject_info["normal_hours"]
        assigned_hours = sum(
            theory_hours_assigned[day][class_name].get(subject_name, 0) for day in range(self.num_days)
        )
        
        # Check if assigned hours are less than normal hours
        if assigned_hours < normal_hours:
            hours_to_allocate = normal_hours - assigned_hours
            
            # print(f"Allocating hours for {subject_name} in {class_name}\n")
            # print(f"Normal hours: {normal_hours}, Assigned hours: {assigned_hours}, Hours to allocate: {hours_to_allocate}\n")
            
            # # Get free hours for teachers and class
            teachers_for_subject = subject_info["teacher_incharge"]
            free_teacher_hours = self.get_free_teacher_hours(teacher_schedule, teachers_for_subject)
            free_class_hours = self.get_free_class_hours(timetable, class_name)

            # Attempt to allocate the missing hours
            for day in range(self.num_days):
                if hours_to_allocate <= 0:
                    # print(f"All required hours for {subject_name} have been allocated.")
                    break  # Exit if all required hours have been allocated

                total_hours_day = theory_hours_assigned[day][class_name].get(subject_name, 0)

                # # Debugging check: Print if more than 2 hours are already assigned for the day
                if total_hours_day >= 2:
                    # print(f"Skipping day {day} for {subject_name} as 2 hours already assigned.")
                    continue

                # Find the available slots for both teacher and class
                available_slots = free_teacher_hours[day].intersection(free_class_hours[day])

                # Allocate hours from available slots
                for hour in available_slots:
                    # Check if the subject has already been allocated at this time
                    # if subject_name in allocated_subjects[day][class_name]:
                    #     print(f"Skipping day {day}, hour {hour}: {subject_name} already allocated for this class.")
                    #     continue
                    
                    # Assign subject and update records
                    
                    if self.is_teacher_allocated_in_first_two_hours( teachers_for_subject, day, hour):
 
                        self.assign_subject(timetable, day, class_name, hour, subject_name, teachers_for_subject, "NORMAL", teacher_schedule)
                    
                                   
                    theory_hours_assigned[day][class_name][subject_name] = total_hours_day + 1
                    allocated_subjects[day][class_name].add(subject_name)
                    total_hours_day += 1  # Increment day total
                    hours_to_allocate -= 1  # Decrement the hours to allocate
                    
                    # Break if all hours have been allocated
                    if hours_to_allocate <= 0:
                        # print(f"All required hours for {subject_name} have been allocated.")
                        break

                # Stop assigning if all needed hours have been allocated
                if hours_to_allocate <= 0:
                    break

            # Optional: Check if after allocation there are still unallocated hours
            if hours_to_allocate > 0:
                print(f"Warning: Unable to allocate all required hours for {subject_name} in {class_name}. Remaining hours to allocate: {hours_to_allocate}.")
        else:
            print(f"{subject_name} in {class_name} already has all required hours allocated.")

    
        
    def get_free_teacher_hours(self, teacher_schedule, teachers_for_subject):
        free_teacher_hours = {}

        for day in range(self.num_days):
            free_hours = set()  # Set to hold all free hours for the day
            for hour in range(self.num_hours):
                for teacher in teachers_for_subject:
                    if teacher not in self.teacher_schedule[day][hour]:  # Teacher is free
                        free_hours.add(hour)  # Add the hour to the set of free hours
                        break  # Exit the loop as we found at least one free teacher for this hour
            free_teacher_hours[day] = free_hours  # Assign the set of free hours for the day

        return free_teacher_hours

    def get_free_class_hours(self, timetable, class_name):
        free_class_hours = {}

        for day in range(self.num_days):
            free_hours = set()  # Set to hold all free hours for the class
            for hour in range(self.num_hours):
                if len(timetable[day][class_name][hour]) == 0:  # Class is free at this hour
                    free_hours.add(hour)  # Add the hour to the set of free hours
            free_class_hours[day] = free_hours  # Assign the set of free hours for the day

        return free_class_hours


    



    def assign_subject(self, timetable, day, class_name, hour, subject_name, teachers_for_subject, category, teacher_schedule):
        if category == "GROUP TEACHING":
            assigned_teachers = []
            for teacher in teachers_for_subject:
                if teacher not in self.teacher_schedule[day][hour]:  # Ensure no double booking
                    assigned_teachers.append(teacher)
                    self.teacher_schedule[day][hour].add(teacher)  # Mark teacher as assigned for this hour

            if assigned_teachers:
                timetable[day][class_name][hour].append((subject_name, assigned_teachers))
                return True  # Return True indicating the subject was assigned
            else:
                print(f"No available teachers for group teaching of {subject_name} in {class_name} on day {day}, hour {hour}.")
                return False  # Return False indicating assignment failed

        elif category == "NORMAL":
            if not any(subject[0] == subject_name for subject in timetable[day][class_name][hour]):
                teacher = random.choice(teachers_for_subject)
                if teacher not in self.teacher_schedule[day][hour]:  # Ensure no double booking
                    timetable[day][class_name][hour].append((subject_name, [teacher]))
                    self.teacher_schedule[day][hour].add(teacher)  # Mark teacher as assigned for this hour
                    return True  # Return True indicating the subject was assigned
                else:
                    print(f"Teacher {teacher} is already booked for {subject_name} in {class_name} on day {day}, hour {hour}.")
                    return False  # Return False indicating assignment failed
            else:
                print(f"{subject_name} is already assigned in {class_name} on day {day}, hour {hour}.")
                return False  # Return False indicating assignment failed



    def get_two_different_available_labs(self, day, hour):
        available_labs = []
        
        # Iterate over lab names to find two available labs
        for lab_name in self.lab_timetables.keys():
            if len(self.lab_timetables[lab_name][day][hour]) == 0 and len(self.lab_timetables[lab_name][day][hour + 1]) == 0:
                available_labs.append(lab_name)
            
            # If two labs are found, return them
            if len(available_labs) == 2:
                return available_labs[0], available_labs[1]
        
        # Return None if fewer than two labs are available
        return None

    def get_available_lab(self, day, hour):
        # Check available lab rooms for the specified day and hour
        for lab_name in self.lab_timetables.keys():
            if len(self.lab_timetables[lab_name][day][hour]) == 0 and len(self.lab_timetables[lab_name][day][hour + 1]) == 0:
                return lab_name
        # return "BSc Lab"
        

    def is_lab_available(self, lab_name, day, hour):
        
        return len(self.lab_timetables[lab_name][day][hour]) == 0

    def assign_lab(self, timetable, day, class_name, hour, subject_name, teachers_for_subject, lab_venue, allocated_subjects, teacher_schedule):
        # Store subject, teachers, and lab venue
        # self.lab_count[class_name][subject_name]=len(teachers_for_subject)
        
        timetable[day][class_name][hour].append((subject_name, teachers_for_subject, lab_venue))  # Store all teachers and lab venue
        timetable[day][class_name][hour + 1].append((subject_name, teachers_for_subject, lab_venue))  # Assign for the next hour
        self.lab_timetables[lab_venue][day][hour].append((subject_name, teachers_for_subject))
        self.lab_timetables[lab_venue][day][hour + 1].append((subject_name, teachers_for_subject))
        allocated_subjects[day][class_name].add(subject_name)
        for teacher in teachers_for_subject:
            self.teacher_schedule[day][hour].add(teacher)
            self.teacher_schedule[day][hour + 1].add(teacher)

        return timetable
    
    def get_free_hours_timetable(self, timetable):
        free_hours = set()  # Set to store free hours

        for day in range(self.num_days):
            for class_name in self.classes:
                for hour in range(self.num_hours):
                    # If the hour is free (i.e., no subjects assigned)
                    if len(timetable[day][class_name][hour]) == 0:
                        free_hours.add((day, hour, class_name))  # Add the free slot to the set

        return free_hours
    


    def create_class_dataframes(self, timetable):
        all_class_dfs = {}
        days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]

        for class_name in self.ct:
            data = {day: [] for day in days}

            for day in range(self.num_days):
                start_hour=0 if class_name in m  else 2
                end_hour=7 if class_name in m else 9
                
                for hour in range(start_hour, end_hour):
                    subjects = timetable[day][class_name][hour]
                    if subjects:
                        subject_info = []
                        for subject in subjects:
                            subject_short_name = shortsub.get(subject[0], subject[0])  # Short form of subject
                            teachers = ', '.join([short_teachers.get(teacher.strip(), teacher) for teacher in subject[1]])  # Short form of teachers
                            if subject_short_name == "BLOCKED":
                                subject_info.append(f" NA")
                            elif len(subject) == 3:  # Lab
                                subject_info.append(f" {subject_short_name} (Lab - {subject[2]} )")
                            else:  # Theory
                                subject_info.append(f" {subject_short_name} ")

                        # Join subject info with new line for better readability
                        data[days[day]].append('\n'.join(subject_info))
                    else:
                        data[days[day]].append("Free")

            # Create DataFrame for this class
            df = pd.DataFrame(data)
            df.index = self.generate_time_intervals(start_hour, end_hour)

            all_class_dfs[class_name] = df

            # Adjust cell sizes based on content
            self.adjust_dataframe_cells(df)

        return all_class_dfs
    
    def generate_time_intervals(self, start_hour=0, end_hour=9):
        # Define the time intervals as a dictionary
        time_intervals = {
            0: '07:30 AM - 08:15 AM',
            1: '08:15 AM - 09:00 AM',
            2: '09:00 AM - 09:45 AM',
            3: '09:45 AM - 10:45 AM',
            4: '10:45 AM - 11:45 AM',
            5: '11:45 AM - 12:45 PM',
            6: '12:45 PM - 01:45 PM',
            7: '01:45 PM - 02:45 PM',
            8: '02:45 PM - 03:45 PM',
        }

        # Return the values from the dictionary based on the specified range
        return [time_intervals[hour] for hour in range(start_hour, end_hour)]


    # Other methods (create_teacher_dataframes, create_lab_dataframes, etc.) can remain unchanged

    
    

    def create_teacher_dataframes(self, timetable):
        all_teacher_dfs = {}
        days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]

        # Collect all unique teachers from courses
        teachers = set()
        for class_data in self.courses.values():
            for category, subjects in class_data.items():
                for subject_info in subjects.values():
                    teachers.update(subject_info.get("teacher_incharge", []))
                    
        for d in range(self.num_days):
            for hr in range(self.num_hours):
                teachers.update(self.teacher_schedule[d][hr])

        teachers = teachers.intersection(set(short_teachers.keys()))
       
       

        # Initialize a timetable for each teacher
        for teacher in teachers:
            data = {day: [] for day in days}

            for day in range(self.num_days):
                for hour in range(self.num_hours):
                    subject_info = []
                    for class_name in self.classes:
                        subjects = timetable[day][class_name][hour]
                        
                        for subject in subjects:
                            
                            if teacher in subject[1]:
                                # print(class_name,subject[0],subject[1])# If teacher is assigned
                                
                                subject_short_name = shortsub.get(subject[0], subject[0])
                                if len(subject)>=3:# Short form of subject
                                    subject_info.append(f"{subject_short_name} ({class_name}) {subject[2]}")
                                else:
                                    subject_info.append(f"{subject_short_name} ({class_name})")

                    # Join subject info with new line for better readability
                    data[days[day]].append('\n'.join(subject_info) if subject_info else "--")

            # Create DataFrame for this teacher
            df = pd.DataFrame(data)
            df.index = self.generate_time_intervals()
            all_teacher_dfs[teacher] = df

            # Adjust cell sizes based on content
            self.adjust_dataframe_cells(df)

        return all_teacher_dfs

    def create_lab_dataframes(self, timetable):
        all_lab_dfs = {}
        days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]

        for lab in self.lab_timetables.keys():
            data = {day: [] for day in days}

            for day in range(self.num_days):
                for hour in range(self.num_hours):
                    subjects_info = []
                    for class_name in self.classes:
                        subjects = timetable[day][class_name][hour]
                        for subject in subjects:
                            if isinstance(subject, tuple) and len(subject) == 3:  # Lab subject
                                subject_name = subject[0]
                                teachers_for_subject = subject[1]
                                lab_venue = subject[2]
                                subject_short_name = shortsub.get(subject_name, subject_name)

                                # Check if current lab matches the subject's lab
                                if lab_venue == lab:
                                    teachers = ', '.join([short_teachers.get(teacher.strip(), teacher) for teacher in teachers_for_subject])
                                    subjects_info.append(f"{subject_short_name} (Lab - {lab} by {teachers}) in {class_name}")

                    data[days[day]].append('\n'.join(subjects_info) if subjects_info else "Free")

            # Create DataFrame for this lab
            df = pd.DataFrame(data)
            df.index = self.generate_time_intervals()
            all_lab_dfs[lab] = df

            # Adjust cell sizes based on content
            self.adjust_dataframe_cells(df)

        return all_lab_dfs
    
    
    def creat_extra_class_dataframes(self, timetable):
        all_class_dfs = {}
        days = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]

        for class_name in self.ct:
            data = {day: [] for day in days}

            for day in range(self.num_days):
                start_hour=0 if class_name in m  else 3
                end_hour=7 if class_name in m else 9
                for hour in range(start_hour,end_hour):
                    subjects = timetable[day][class_name][hour]
                    if subjects:
                        subject_info = []
                        for subject in subjects:
                            subject_short_name = subject[0] # Short form of subject
                            teachers = ', '.join([ teacher for teacher in subject[1]])  # Short form of teachers
                            if subject[0] in cc[class_name]:
                                code=cc[class_name][subject[0]]
                            if subject_short_name=="BLOCKED":
                                subject_info.append(f" NA")
                            elif len(subject) == 3:  # Lab
                                subject_info.append(f" ({code}) (Lab - {subject[2]} by {teachers})")
                                code=""
                            else:  # Theory
                                subject_info.append(f" ({code}) ({subject[0]}) by {teachers}")
                                code=""

                        # Join subject info with new line for better readability
                        data[days[day]].append('\n'.join(subject_info)) 
                    else:
                        data[days[day]].append("Free")

            # Create DataFrame for this class
            df = pd.DataFrame(data) 
            df.index = self.generate_time_intervals(start_hour,end_hour)

            all_class_dfs[class_name] = df

            # Adjust cell sizes based on content
            self.adjust_dataframe_cells(df)

        return all_class_dfs
    
    
    


    def adjust_dataframe_cells(self, df):
        """ Adjust the cell size based on content for better display. """
        # Maximum length for columns based on content
        max_length = {col: max(df[col].astype(str).map(len).max(), len(col)) for col in df.columns}

        # Set a minimum width (for example, 35 characters) and maximize width
        for column in df.columns:
            # max_length[column] = max(max_length[column], 0)  # Minimum width of 35 characters

            # Left justify each cell content
            df[column] = df[column].apply(lambda x: x.ljust(max_length[column]))

        # Also adjust the index (hour labels)
        max_index_length = max(df.index.astype(str).map(len).max(), len("Hour N"))
        df.index = df.index.map(lambda x: x.ljust(max_index_length))  # Left justify index labels


    def style_and_transpose_excel_sheet(self, file_name, output_file_name):
        # Load the workbook
        workbook = openpyxl.load_workbook(file_name)

        # Define styles
        header_font = Font(bold=True, color="FFFFFF")  # White text
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")  # Blue background
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # Iterate over each sheet in the workbook
        for sheet in workbook.worksheets:
            # Read the current content of the sheet
            content = []
            for row in sheet.iter_rows(values_only=True):
                content.append(list(row))
            
            # Transpose the table
            transposed_content = list(map(list, zip(*content)))

            # Clear the sheet and reapply transposed content
            sheet.delete_rows(1, sheet.max_row)
            for row_idx, row_data in enumerate(transposed_content, start=1):
                for col_idx, cell_value in enumerate(row_data, start=1):
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    cell.value = cell_value

            # Adjust column widths and style headers
            for column in sheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)  # Updated to use get_column_letter

                for cell in column:
                    # Adjust maximum length for column width
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass

                    # Apply borders to each cell
                    cell.border = border

                adjusted_width = (max_length + 2)  # Add a little extra space for padding
                sheet.column_dimensions[column_letter].width = adjusted_width  # Set the width

            # Style the header row (assumes headers are in the first row)
            for cell in sheet[1]:  # The first row (index 1)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border

        # Save the modified workbook
        workbook.save(output_file_name)  # Save it with the specified output filename





    


    def apply_styles(self,cell, font=None, fill=None, border=None, alignment=None):
        if font:
            cell.font = font
        if fill:
            cell.fill = fill
        if border:
            cell.border = border
        if alignment:
            cell.alignment = alignment

    # Helper function to auto-adjust column width and ensure wrapping
    def auto_adjust_column_width(self,sheet):
        wrap_alignment = Alignment(wrap_text=True)

        for col in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(col[0].column)

            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
                # Ensure text wrapping is enabled
                cell.alignment = wrap_alignment
            
        #     # Adjust column width based on maximum content length
        #     adjusted_width = (max_length + 2) * 1.2
        #     sheet.column_dimensions[column_letter].width = adjusted_width

    # Function to merge sheets into one, with automatic wrapping and cell sizing
    def merge_sheets_to_one_wrapped(self,file_path, output_file_path, space_between_sheets=2):
        # Load the Excel workbook
        x=self.calculate_teacher_hours()
        wb = openpyxl.load_workbook(file_path)
        
        # Create a new workbook for the output
        new_wb = Workbook()
        new_ws = new_wb.active
        new_ws.title = "TimeTable"
        
        # Define the styles
        header_font = Font(bold=True, size=14)
        header_alignment = Alignment(horizontal="center", vertical="center")
        header_fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        # Variable to keep track of the current row in the output sheet
        current_row = 1
        
        # Loop through each sheet in the workbook
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]  # Access the worksheet by name
            
            # Write the sheet name at the top of the content with styling
            new_ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=ws.max_column)
            header_cell = new_ws.cell(row=current_row, column=1)
            h = x.get(sheet_name, "")
            if h != "":
                h = "\t Hours:" + str(h)

            header_cell.value = sheet_name+h
            self.apply_styles(header_cell, font=header_font, alignment=header_alignment, fill=header_fill)
            current_row += 1  # Move to the next row
            
            # Copy content from the current sheet to the new sheet with wrapping and borders
            for row in ws.iter_rows(values_only=True):
                for col_idx, cell_value in enumerate(row, start=1):
                    new_cell = new_ws.cell(row=current_row, column=col_idx)
                    new_cell.value = cell_value
                    self.apply_styles(new_cell, border=thin_border)
                current_row += 1  # Move to the next row

            # Add space between sheets
            current_row += space_between_sheets
        
        # Adjust column widths and enable wrapping for all content
        self.auto_adjust_column_width(new_ws)

        # Save the new workbook
        new_wb.save(output_file_path)
        # print(f"All sheets have been merged with wrapping into {output_file_path}")
        

    def coursett(self):
        # Create a new workbook
        wb = Workbook()

        # Iterate through the courses dictionary
        for class_name, categories in courses.items():
            data = []  # Reset data for each class
            class_written = False  # Flag to track if the class name has been written

            for category, subjects in categories.items():
                for subject_name, details in subjects.items():
                    if 'subject_code' in details:
                        # Handle regular subjects
                        data.append({
                            "Class Name": class_name if not class_written else "",
                            "Subject Code": details['subject_code'],
                            "Subject Name": shortsub[subject_name],
                            "Teacher Incharge": ', '.join(details['teacher_incharge'])  # Join the list of teachers
                        })
                        class_written = True
                        
                        # Handle lab subjects if applicable
                        if "lab_hours" in details and details["lab_hours"] > 0:
                            flat_teachers = [teacher for sublist in self.lab_teachers[class_name][subject_name] for teacher in sublist]
                            teachers_in_charge = ", ".join(flat_teachers) if flat_teachers else "No teacher assigned"

                            data.append({
                                "Class Name": "",  # No class name for lab row
                                "Subject Code": details['subject_code'],
                                "Subject Name": shortsub[subject_name]+" Lab",
                                "Teacher Incharge": teachers_in_charge
                            })

            # Add two empty rows after each class for spacing
            data.append({"Class Name": "", "Subject Code": "", "Subject Name": "", "Teacher Incharge": ""})
            data.append({"Class Name": "", "Subject Code": "", "Subject Name": "", "Teacher Incharge": ""})

            # Create a DataFrame from the data list for the current class
            df = pd.DataFrame(data)

            # Create a new sheet for the current class with the class name as the title
            ws = wb.create_sheet(title="teachers "+class_name)

            # Write the DataFrame to the new sheet
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)

            # Adjust column sizes based on content
            for column_cells in ws.columns:
                max_length = 0
                column = column_cells[0].column_letter  # Get the column letter
                for cell in column_cells:
                    try:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))  # Get the max length of content in column
                    except Exception as e:
                        print(f"Error processing cell value: {cell.value}. Error: {e}")
                adjusted_width = max_length + 4  # Add some padding
                ws.column_dimensions[column].width = adjusted_width

        # Remove the default sheet created by openpyxl, if it exists
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

        # Save the modified workbook
        wb.save('courses_timetable.xlsx')

        





if __name__ == "__main__":
    timetable_generator = TimetableGenerator(courses)
    timetable = timetable_generator.generate_timetable()

    # Create DataFrames for each class
    class_dataframes = timetable_generator.create_class_dataframes(timetable)

    # # Create DataFrames for each teacher
    teacher_dataframes = timetable_generator.create_teacher_dataframes(timetable)

    lab_dataframes = timetable_generator.create_lab_dataframes(timetable)
    
    extra=timetable_generator.creat_extra_class_dataframes(timetable)
    timetable_generator.coursett()
    

    # Save each class DataFrame to separate sheets in an Excel file
    with pd.ExcelWriter("timetable.xlsx") as writer:
        for class_name, df in class_dataframes.items():
            df.to_excel(writer, sheet_name=class_name, index=True)
    
    with pd.ExcelWriter("extra_timetable.xlsx") as writer:
        for class_name, df in extra.items():
            df.to_excel(writer, sheet_name=class_name, index=True)
            
    with pd.ExcelWriter("Lab_timetable.xlsx") as writer:    
        for lab_name, df in lab_dataframes.items():
            df.to_excel(writer, sheet_name=lab_name, index=True)
    

    # Save teacher DataFrames to separate sheets
    with pd.ExcelWriter("teachertimetable.xlsx") as writer:
        for teacher_name, df in teacher_dataframes.items():
            df.to_excel(writer, sheet_name=teacher_name, index=True)
    
    timetable_generator.style_and_transpose_excel_sheet('timetable.xlsx', 'timetable.xlsx')
    timetable_generator.style_and_transpose_excel_sheet('extra_timetable.xlsx', 'extra_timetable.xlsx')
    timetable_generator.style_and_transpose_excel_sheet('Lab_timetable.xlsx', 'Lab_timetable.xlsx')
    timetable_generator.style_and_transpose_excel_sheet('teachertimetable.xlsx', 'teachertimetable.xlsx')
    
    timetable_generator.merge_sheets_to_one_wrapped("timetable.xlsx", "timetable.xlsx", 4)
    timetable_generator.merge_sheets_to_one_wrapped("courses_timetable.xlsx", "courses_timetable.xlsx", 4)
    
    timetable_generator.merge_sheets_to_one_wrapped("extra_timetable.xlsx", "extra_timetable.xlsx", 4)
    timetable_generator.merge_sheets_to_one_wrapped("Lab_timetable.xlsx", "Lab_timetable.xlsx", 4)
    timetable_generator.merge_sheets_to_one_wrapped("teachertimetable.xlsx", "teachertimetable.xlsx", 4)
    
    

    # Define the file paths for both Excel files
    import pandas as pd

# Define the file paths for both Excel files
    file1 = 'timetable.xlsx'
    file2 = 'courses_timetable.xlsx'

    # Create a writer object to write to the merged file
    writer = pd.ExcelWriter('full_tt.xlsx', engine='openpyxl')

    # Load the first file and write all its sheets to the new file
    xls1 = pd.ExcelFile(file1)
    for sheet_name in xls1.sheet_names:
        # Read each sheet into a DataFrame
        df = pd.read_excel(file1, sheet_name=sheet_name)
        # Write each sheet into the merged Excel file
        df.to_excel(writer, sheet_name=sheet_name, index=False)

    # Load the second file and append its sheets to the new file
    xls2 = pd.ExcelFile(file2)
    for sheet_name in xls2.sheet_names:
        # Read each sheet into a DataFrame
        df = pd.read_excel(file2, sheet_name=sheet_name)
        # Write each sheet into the merged Excel file with a suffix to avoid name conflicts
        df.to_excel(writer, sheet_name=sheet_name+"tt" , index=False)

    # Save the merged file
    writer.close()

            
    print("Timetable saved to 'timetable.xlsx'")
    
    
 